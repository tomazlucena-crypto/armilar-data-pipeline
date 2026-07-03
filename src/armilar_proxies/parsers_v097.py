"""Deterministic parsers and normalized-observation contract for ARMILAR v0.9.7."""
from __future__ import annotations

import csv
import io
import json
import re
from typing import Any, Callable, Iterable, Iterator, Mapping, Sequence

from .core_v097 import (
    PARSER_VERSION,
    ProxyRegistryError,
    _period_not_after_retrieval,
    decimal_text,
    normalise_label,
    parse_month,
    parse_quarter,
    parse_week,
)

def _import_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise ProxyRegistryError("openpyxl is required for XLSX proxy sources; install the 'proxies' extra") from exc
    return openpyxl


def _observation(
    *,
    source: Mapping[str, Any],
    series_id: str,
    proxy_domain: str,
    geography: str,
    period: str,
    value: Any,
    unit: str,
    published_at: str | None,
    retrieved_at: str,
    raw_snapshot_id: str,
    source_sha256: str,
    registry_sha256: str,
    row_locator: str,
) -> dict[str, str]:
    return {
        "source_id": source["source_id"],
        "proxy_domain": proxy_domain,
        "series_id": series_id,
        "geography": geography,
        "period": period,
        "frequency": source["frequency"],
        "value": decimal_text(value),
        "unit": unit,
        "published_at": published_at or "",
        "publication_time_status": source["publication_time_status"],
        "retrieved_at": retrieved_at,
        "raw_snapshot_id": raw_snapshot_id,
        "source_sha256": source_sha256,
        "registry_sha256": registry_sha256,
        "row_locator": row_locator,
        "parser_version": PARSER_VERSION,
        "direct_index_use_allowed": "false",
        "arm_l_use_allowed": "false",
        "model_training_allowed": "false",
    }


def parse_fao_ffpi_csv(
    payload: bytes,
    *,
    source: Mapping[str, Any],
    published_at: str | None,
    retrieved_at: str,
    raw_snapshot_id: str,
    source_sha256: str,
    registry_sha256: str,
) -> list[dict[str, str]]:
    text = payload.decode("utf-8-sig")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    header_index = None
    header: list[str] = []
    for index, row in enumerate(rows):
        labels = [normalise_label(cell) for cell in row]
        if any(label in {"date", "month", "period"} for label in labels) and any("food" in label and "price" in label for label in labels):
            header_index = index
            header = labels
            break
    if header_index is None:
        raise ProxyRegistryError("FAO CSV header row was not found")
    date_column = next(i for i, label in enumerate(header) if label in {"date", "month", "period"})
    aliases = {
        "FAO_FFPI": ("FOOD", ["food price index", "food price"]),
        "FAO_MEAT": ("FOOD", ["meat price index", "meat"]),
        "FAO_DAIRY": ("FOOD", ["dairy price index", "dairy"]),
        "FAO_CEREALS": ("FOOD", ["cereals price index", "cereal price index", "cereals"]),
        "FAO_OILS": ("FOOD", ["oils price index", "vegetable oil price index", "oils"]),
        "FAO_SUGAR": ("FOOD", ["sugar price index", "sugar"]),
    }
    selected: dict[int, tuple[str, str]] = {}
    for series_id, (domain, candidates) in aliases.items():
        for i, label in enumerate(header):
            if label in candidates:
                selected[i] = (series_id, domain)
                break
    if "FAO_FFPI" not in {item[0] for item in selected.values()}:
        raise ProxyRegistryError("FAO Food Price Index column was not found")
    output: list[dict[str, str]] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if date_column >= len(row) or not str(row[date_column]).strip():
            continue
        try:
            period = parse_month(row[date_column])
        except ProxyRegistryError:
            continue
        for column, (series_id, domain) in selected.items():
            if column >= len(row) or not str(row[column]).strip():
                continue
            try:
                output.append(
                    _observation(
                        source=source,
                        series_id=series_id,
                        proxy_domain=domain,
                        geography="WORLD",
                        period=period,
                        value=row[column],
                        unit="INDEX_2014_2016_100",
                        published_at=published_at,
                        retrieved_at=retrieved_at,
                        raw_snapshot_id=raw_snapshot_id,
                        source_sha256=source_sha256,
                        registry_sha256=registry_sha256,
                        row_locator=f"csv:{row_number}:{column + 1}",
                    )
                )
            except ProxyRegistryError:
                continue
    if not output:
        raise ProxyRegistryError("FAO parser produced no observations")
    return output


def _find_header_row(rows: Sequence[Sequence[Any]], *, date_aliases: set[str], minimum_nonempty: int = 3) -> tuple[int, list[str]]:
    for index, row in enumerate(rows):
        labels = [normalise_label(cell) for cell in row]
        nonempty = sum(bool(label) for label in labels)
        if nonempty >= minimum_nonempty and any(label in date_aliases for label in labels):
            return index, labels
    raise ProxyRegistryError("tabular date header row was not found")


def parse_world_bank_pink_sheet_xlsx(
    payload: bytes,
    *,
    source: Mapping[str, Any],
    published_at: str | None,
    retrieved_at: str,
    raw_snapshot_id: str,
    source_sha256: str,
    registry_sha256: str,
) -> list[dict[str, str]]:
    openpyxl = _import_openpyxl()
    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = None
    for candidate in workbook.worksheets:
        if "monthly" in normalise_label(candidate.title) and "price" in normalise_label(candidate.title):
            sheet = candidate
            break
    if sheet is None:
        sheet = workbook.worksheets[0]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    try:
        header_index, labels = _find_header_row(rows, date_aliases={"date", "month", "period"})
        date_column = next(i for i, label in enumerate(labels) if label in {"date", "month", "period"})
    except ProxyRegistryError:
        header_index = None
        labels = []
        date_column = 0
        for data_index, row in enumerate(rows):
            if not row:
                continue
            try:
                parse_month(row[0])
            except ProxyRegistryError:
                continue
            if data_index < 2:
                continue
            header_index = data_index - 2
            labels = [normalise_label(cell) for cell in rows[header_index]]
            if not labels or not labels[0]:
                labels = ["date", *labels[1:]]
            break
        if header_index is None or not labels:
            raise
    selections = source["series_selection"]
    selected: dict[int, dict[str, str]] = {}
    for output_id, spec in selections.items():
        aliases = {normalise_label(item) for item in spec["aliases"]}
        for i, label in enumerate(labels):
            if label in aliases:
                selected[i] = {"series_id": output_id, "domain": spec["proxy_domain"], "unit": spec["unit"]}
                break
    if not selected:
        raise ProxyRegistryError("World Bank selected Pink Sheet columns were not found")
    output: list[dict[str, str]] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if date_column >= len(row) or row[date_column] is None:
            continue
        try:
            period = parse_month(row[date_column])
        except ProxyRegistryError:
            continue
        for column, spec in selected.items():
            if column >= len(row) or row[column] is None:
                continue
            try:
                output.append(
                    _observation(
                        source=source,
                        series_id=spec["series_id"],
                        proxy_domain=spec["domain"],
                        geography="WORLD",
                        period=period,
                        value=row[column],
                        unit=spec["unit"],
                        published_at=published_at,
                        retrieved_at=retrieved_at,
                        raw_snapshot_id=raw_snapshot_id,
                        source_sha256=source_sha256,
                        registry_sha256=registry_sha256,
                        row_locator=f"xlsx:{sheet.title}:{row_number}:{column + 1}",
                    )
                )
            except ProxyRegistryError:
                continue
    if not output:
        raise ProxyRegistryError("World Bank parser produced no observations")
    return output


def _sheet_product_id(title: str) -> str:
    label = normalise_label(title)
    if "diesel" in label or "gasoil" in label:
        return "DIESEL"
    if "95" in label or "eurosuper" in label or "gasoline" in label or "petrol" in label:
        return "PETROL_95"
    if "heating" in label:
        return "HEATING_OIL"
    if "lpg" in label:
        return "LPG"
    return re.sub(r"[^A-Z0-9]+", "_", title.upper()).strip("_") or "UNSPECIFIED"


def parse_ec_oil_bulletin_xlsx(
    payload: bytes,
    *,
    source: Mapping[str, Any],
    published_at: str | None,
    retrieved_at: str,
    raw_snapshot_id: str,
    source_sha256: str,
    registry_sha256: str,
) -> list[dict[str, str]]:
    openpyxl = _import_openpyxl()
    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    output: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    country_aliases = {"country", "member state", "memberstate", "geo", "geography"}
    date_aliases = {"date", "week", "observation date", "reference date"}
    price_aliases = {"price", "value", "eur 1000 l", "eur per 1000 l", "euro 1000 litres"}
    product_aliases = {"product", "fuel", "petroleum product"}
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        if not rows:
            continue
        # Long format: date, country, product, price.
        header_candidate = None
        for index, row in enumerate(rows[:60]):
            labels = [normalise_label(cell) for cell in row]
            if any(label in date_aliases for label in labels) and any(label in country_aliases for label in labels) and any(label in price_aliases for label in labels):
                header_candidate = (index, labels)
                break
        if header_candidate:
            header_index, labels = header_candidate
            date_col = next(i for i, label in enumerate(labels) if label in date_aliases)
            country_col = next(i for i, label in enumerate(labels) if label in country_aliases)
            price_col = next(i for i, label in enumerate(labels) if label in price_aliases)
            product_col = next((i for i, label in enumerate(labels) if label in product_aliases), None)
            unit_col = next((i for i, label in enumerate(labels) if label in {"unit", "units"}), None)
            for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                if max(date_col, country_col, price_col) >= len(row):
                    continue
                if row[date_col] is None or row[country_col] is None or row[price_col] is None:
                    continue
                try:
                    period = parse_week(row[date_col])
                    product = str(row[product_col]).strip() if product_col is not None and product_col < len(row) and row[product_col] else _sheet_product_id(sheet.title)
                    unit = str(row[unit_col]).strip() if unit_col is not None and unit_col < len(row) and row[unit_col] else "EUR_PER_1000_LITRES"
                    output.append(
                        _observation(
                            source=source,
                            series_id=f"EC_OIL_{_sheet_product_id(product)}",
                            proxy_domain="TRANSPORT" if _sheet_product_id(product) in {"DIESEL", "PETROL_95", "LPG"} else "FUELS",
                            geography=str(row[country_col]).strip().upper(),
                            period=period,
                            value=row[price_col],
                            unit=unit,
                            published_at=published_at,
                            retrieved_at=retrieved_at,
                            raw_snapshot_id=raw_snapshot_id,
                            source_sha256=source_sha256,
                            registry_sha256=registry_sha256,
                            row_locator=f"xlsx:{sheet.title}:{row_number}:{price_col + 1}",
                        )
                    )
                except ProxyRegistryError:
                    continue
                identity = (output[-1]["source_id"], output[-1]["series_id"], output[-1]["geography"], output[-1]["period"], output[-1]["unit"])
                if identity in seen:
                    output.pop()
                    continue
                seen.add(identity)
            continue
        # Wide format: date in first identified column; countries across columns.
        try:
            header_index, labels = _find_header_row(rows[:80], date_aliases=date_aliases, minimum_nonempty=4)
        except ProxyRegistryError:
            continue
        date_col = next(i for i, label in enumerate(labels) if label in date_aliases)
        product = _sheet_product_id(sheet.title)
        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if date_col >= len(row) or row[date_col] is None:
                continue
            try:
                period = parse_week(row[date_col])
            except ProxyRegistryError:
                continue
            for column, geography_label in enumerate(rows[header_index]):
                if column == date_col or geography_label is None or column >= len(row) or row[column] is None:
                    continue
                geography = str(geography_label).strip().upper()
                if not geography or normalise_label(geography) in {"unit", "product", "notes"}:
                    continue
                try:
                    output.append(
                        _observation(
                            source=source,
                            series_id=f"EC_OIL_{product}",
                            proxy_domain="TRANSPORT" if product in {"DIESEL", "PETROL_95", "LPG"} else "FUELS",
                            geography=geography,
                            period=period,
                            value=row[column],
                            unit="EUR_PER_1000_LITRES",
                            published_at=published_at,
                            retrieved_at=retrieved_at,
                            raw_snapshot_id=raw_snapshot_id,
                            source_sha256=source_sha256,
                            registry_sha256=registry_sha256,
                            row_locator=f"xlsx:{sheet.title}:{row_number}:{column + 1}",
                        )
                    )
                except ProxyRegistryError:
                    continue
                identity = (output[-1]["source_id"], output[-1]["series_id"], output[-1]["geography"], output[-1]["period"], output[-1]["unit"])
                if identity in seen:
                    output.pop()
                    continue
                seen.add(identity)
    if not output:
        raise ProxyRegistryError("European Commission Oil Bulletin parser produced no observations")
    return output


def _jsonstat_category_values(dimension: Mapping[str, Any], name: str) -> list[str]:
    block = dimension.get(name)
    if not isinstance(block, dict):
        raise ProxyRegistryError(f"Eurostat dimension missing: {name}")
    category = block.get("category")
    if not isinstance(category, dict):
        raise ProxyRegistryError(f"Eurostat category missing: {name}")
    index = category.get("index")
    if isinstance(index, list):
        return [str(value) for value in index]
    if isinstance(index, dict):
        return [key for key, _ in sorted(index.items(), key=lambda item: int(item[1]))]
    raise ProxyRegistryError(f"Eurostat category index invalid: {name}")


def _cartesian_indices(sizes: Sequence[int]) -> Iterator[tuple[int, ...]]:
    if not sizes:
        yield ()
        return
    def walk(prefix: tuple[int, ...], depth: int) -> Iterator[tuple[int, ...]]:
        if depth == len(sizes):
            yield prefix
            return
        for value in range(sizes[depth]):
            yield from walk(prefix + (value,), depth + 1)
    yield from walk((), 0)


def parse_eurostat_jsonstat(
    payload: bytes,
    *,
    source: Mapping[str, Any],
    published_at: str | None,
    retrieved_at: str,
    raw_snapshot_id: str,
    source_sha256: str,
    registry_sha256: str,
) -> list[dict[str, str]]:
    data = json.loads(payload.decode("utf-8-sig"))
    if not isinstance(data, dict):
        raise ProxyRegistryError("Eurostat JSON-stat root must be an object")
    ids = data.get("id")
    sizes = data.get("size")
    dimension = data.get("dimension")
    values = data.get("value")
    if not isinstance(ids, list) or not isinstance(sizes, list) or len(ids) != len(sizes) or not isinstance(dimension, dict):
        raise ProxyRegistryError("Eurostat JSON-stat dimensions are invalid")
    if not all(isinstance(size, int) and size >= 0 for size in sizes):
        raise ProxyRegistryError("Eurostat JSON-stat sizes are invalid")
    categories = {name: _jsonstat_category_values(dimension, name) for name in ids}
    geo_name = next((name for name in ids if name.lower() in {"geo", "geography"}), None)
    time_name = next((name for name in ids if name.lower() in {"time", "period"}), None)
    if geo_name is None or time_name is None:
        raise ProxyRegistryError("Eurostat OOHPI dataset requires geo and time dimensions")
    value_map: dict[int, Any]
    if isinstance(values, list):
        value_map = {index: value for index, value in enumerate(values) if value is not None}
    elif isinstance(values, dict):
        value_map = {int(index): value for index, value in values.items() if value is not None}
    else:
        raise ProxyRegistryError("Eurostat JSON-stat value block is invalid")
    output: list[dict[str, str]] = []
    multipliers: list[int] = []
    running = 1
    for size in reversed(sizes[1:]):
        running *= size
        multipliers.insert(0, running)
    multipliers.append(1)
    for coordinates in _cartesian_indices(sizes):
        flat_index = sum(coordinates[i] * multipliers[i] for i in range(len(sizes)))
        if flat_index not in value_map:
            continue
        point = {ids[i]: categories[ids[i]][coordinates[i]] for i in range(len(ids))}
        try:
            period = parse_quarter(point[time_name])
        except ProxyRegistryError:
            continue
        geo = point[geo_name]
        unit = point.get("unit", "INDEX")
        item = point.get("expend") or point.get("purchase") or point.get("coicop") or point.get("item") or "TOTAL"
        series_id = f"EUROSTAT_OOHPI_{item}_{unit}"
        try:
            output.append(
                _observation(
                    source=source,
                    series_id=series_id,
                    proxy_domain="HOUSING_OOH_SENSITIVITY",
                    geography=geo,
                    period=period,
                    value=value_map[flat_index],
                    unit=unit,
                    published_at=published_at,
                    retrieved_at=retrieved_at,
                    raw_snapshot_id=raw_snapshot_id,
                    source_sha256=source_sha256,
                    registry_sha256=registry_sha256,
                    row_locator=f"jsonstat:{flat_index}",
                )
            )
        except ProxyRegistryError:
            continue
    if not output:
        raise ProxyRegistryError("Eurostat OOHPI parser produced no observations")
    return output


PARSERS: dict[str, Callable[..., list[dict[str, str]]]] = {
    "fao_ffpi_csv_v1": parse_fao_ffpi_csv,
    "world_bank_pink_sheet_xlsx_v1": parse_world_bank_pink_sheet_xlsx,
    "ec_oil_bulletin_xlsx_v1": parse_ec_oil_bulletin_xlsx,
    "eurostat_jsonstat_v1": parse_eurostat_jsonstat,
}


OBSERVATION_COLUMNS = [
    "source_id",
    "proxy_domain",
    "series_id",
    "geography",
    "period",
    "frequency",
    "value",
    "unit",
    "published_at",
    "publication_time_status",
    "retrieved_at",
    "raw_snapshot_id",
    "source_sha256",
    "registry_sha256",
    "row_locator",
    "parser_version",
    "direct_index_use_allowed",
    "arm_l_use_allowed",
    "model_training_allowed",
]


def sort_observations(rows: Iterable[Mapping[str, str]]) -> list[dict[str, str]]:
    normalised = [dict(row) for row in rows]
    identities: set[tuple[str, str, str, str, str]] = set()
    for row in normalised:
        if set(row) != set(OBSERVATION_COLUMNS):
            raise ProxyRegistryError("normalised observation columns violate the closed contract")
        if row["direct_index_use_allowed"] != "false" or row["arm_l_use_allowed"] != "false" or row["model_training_allowed"] != "false":
            raise ProxyRegistryError("proxy observation attempts to open a prohibited gate")
        if not _period_not_after_retrieval(row["period"], row["frequency"], row["retrieved_at"]):
            raise ProxyRegistryError(f"observation period is after retrieval time: {row['period']}")
        identity = (row["source_id"], row["series_id"], row["geography"], row["period"], row["unit"])
        if identity in identities:
            raise ProxyRegistryError(f"duplicate proxy observation identity: {identity}")
        identities.add(identity)
    return sorted(
        normalised,
        key=lambda row: (
            row["source_id"],
            row["series_id"],
            row["geography"],
            row["period"],
            row["unit"],
            row["row_locator"],
        ),
    )


def csv_bytes(rows: Sequence[Mapping[str, str]]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=OBSERVATION_COLUMNS, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: row[column] for column in OBSERVATION_COLUMNS})
    return stream.getvalue().encode("utf-8")
